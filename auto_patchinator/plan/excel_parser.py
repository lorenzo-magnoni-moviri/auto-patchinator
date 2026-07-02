"""Reads the wave Excel's 'Plan' sheet into a flat, typed list of rows.

Column names vary across waves (English vs Italian naming, suffix-numbered variants):
  Step       : 'Step'       | 'Step2'
  Dependancy : 'Dependancy' | 'Dipendenza' | 'Dipendenza2'
  All other required columns are stable across waves.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Sequence

import openpyxl

# Maps canonical name -> accepted aliases (first match in header wins).
_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "Step":             ("Step", "Step2"),
    "Dependancy":       ("Dependancy", "Dipendenza", "Dipendenza2"),
    "Applicativo":      ("Applicativo",),
    "Nome":             ("Nome",),
    "Gruppo_referente": ("Gruppo_referente",),
    "Stato":            ("Stato",),
    "Tipo":             ("Tipo",),
    "StartDate_Prevista": ("StartDate_Prevista",),
    "EndDate_Prevista": ("EndDate_Prevista",),
    "Note":             ("Note",),
}


@dataclass(frozen=True)
class RawStep:
    step: int
    dependencies: tuple[int, ...]
    applicativo: str | None
    nome: str
    gruppo_referente: str | None
    stato: str | None
    tipo: str | None
    start_date: datetime | None
    end_date: datetime | None
    note: str | None


def _resolve_columns(
    header: Sequence,
    aliases: dict[str, tuple[str, ...]],
) -> dict[str, int]:
    """Return {canonical_name: column_index} for every alias group.

    Raises ValueError listing every canonical name whose aliases are all absent.
    """
    idx: dict[str, int] = {}
    missing: list[str] = []
    for canonical, candidates in aliases.items():
        for name in candidates:
            if name in header:
                idx[canonical] = list(header).index(name)
                break
        else:
            missing.append(f"{canonical!r} (tried: {', '.join(candidates)})")
    if missing:
        raise ValueError(
            f"Plan sheet header {list(header)!r} is missing required columns: "
            + "; ".join(missing)
        )
    return idx


def _parse_dependencies(value) -> tuple[int, ...]:
    if value is None:
        return ()
    if isinstance(value, int):
        return (value,)
    if isinstance(value, float):
        return (int(value),)
    if isinstance(value, str):
        result = []
        for tok in value.split():
            tok = tok.strip()
            if tok:
                try:
                    result.append(int(tok))
                except ValueError:
                    pass  # ignore non-numeric tokens ('Pre', 'task', label refs)
        return tuple(result)
    raise ValueError(f"unexpected Dependancy value: {value!r}")


def load_plan_sheet(path: str | Path, sheet_name: str = "Plan") -> list[RawStep]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"{path}: no sheet named {sheet_name!r}, found {wb.sheetnames}")
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"{path}: sheet {sheet_name!r} is empty")

    idx = _resolve_columns(rows[0], _COLUMN_ALIASES)

    steps: list[RawStep] = []
    for row in rows[1:]:
        raw_step = row[idx["Step"]]
        if raw_step is None:
            continue
        try:
            step_num = int(raw_step)
        except (ValueError, TypeError):
            continue  # skip header-like rows: 'Pre task 1', section labels, etc.
        steps.append(
            RawStep(
                step=step_num,
                dependencies=_parse_dependencies(row[idx["Dependancy"]]),
                applicativo=row[idx["Applicativo"]],
                nome=row[idx["Nome"]] or "",
                gruppo_referente=row[idx["Gruppo_referente"]],
                stato=row[idx["Stato"]],
                tipo=row[idx["Tipo"]],
                start_date=row[idx["StartDate_Prevista"]],
                end_date=row[idx["EndDate_Prevista"]],
                note=row[idx["Note"]],
            )
        )
    return steps
