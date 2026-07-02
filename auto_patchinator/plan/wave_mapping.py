"""Per-wave 'Group N' -> hostnames mapping, read directly from the Excel host sheet.

The 'no-IT' host sheet in each wave's Excel lists all non-IT hosts across every team.
We extract only the groups/hosts that belong to us by cross-referencing with the static
inventory: any host in the sheet that also appears in hosts.yaml is ours.

Sheet and column names vary across waves:
  Sheet name  : 'List Host NO IT' | 'Host list_NO IT_*' | 'Host No IT' | ...
                -> found by case-insensitive substring match on 'NO IT' or 'No IT'
  Group column: 'Group' | 'Groups'
  Hostname    : 'Hostname' (bare, preferred) | stripped from 'Computer' (FQDN, fallback)

Sheet layout quirk: row 0 is a total-count cell, row 1 is the actual column header.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from auto_patchinator.config.inventory import Inventory
from auto_patchinator.plan.action_mapping import TeamStep

HOST_SHEET_NAME = "List Host NO IT"

# Substrings used to locate the host sheet when the exact name isn't found.
_HOST_SHEET_MARKERS = ("NO IT", "No IT")

_HEADER_ROW_INDEX = 1  # row 0 is a count cell; row 1 is the real header

# Alias lists: first match in the header row wins.
_GROUP_ALIASES  = ("Group", "Groups")
_HOST_ALIASES   = ("Hostname",)   # 'Computer' (FQDN) is the fallback, handled separately
_COMPUTER_COL   = "Computer"


def _find_host_sheet(wb: openpyxl.Workbook, hint: str) -> str:
    """Return the sheet name to use for the host→group mapping.

    Tries the exact *hint* first, then falls back to the first sheet whose name
    contains any of the known marker substrings (case-insensitive).
    """
    if hint in wb.sheetnames:
        return hint
    for sheet in wb.sheetnames:
        if any(m.lower() in sheet.lower() for m in _HOST_SHEET_MARKERS):
            return sheet
    raise ValueError(
        f"Could not find the host sheet in {wb}. "
        f"Tried exact name {hint!r} and substring markers {_HOST_SHEET_MARKERS}. "
        f"Available sheets: {wb.sheetnames}. "
        f"Override with --host-sheet."
    )


def _find_col(header: tuple, aliases: tuple[str, ...]) -> int | None:
    for name in aliases:
        if name in header:
            return header.index(name)
    return None


def _bare_hostname(computer: str) -> str:
    """Strip domain suffix from a FQDN to get a bare hostname."""
    return computer.split(".")[0]


def load_wave_mapping_from_excel(
    path: str | Path,
    inventory: Inventory,
    sheet_name: str = HOST_SHEET_NAME,
) -> dict[int, tuple[str, ...]]:
    """Return {group_number: (hostname, ...)} filtered to hosts present in the inventory.

    Handles varying column names and the missing-Hostname fallback automatically.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    actual_sheet = _find_host_sheet(wb, sheet_name)
    ws = wb[actual_sheet]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) <= _HEADER_ROW_INDEX:
        raise ValueError(f"{path}: sheet {actual_sheet!r} has no data rows")

    header = rows[_HEADER_ROW_INDEX]

    group_col = _find_col(header, _GROUP_ALIASES)
    if group_col is None:
        raise ValueError(
            f"{path}: sheet {actual_sheet!r} header has no group column "
            f"(tried: {', '.join(_GROUP_ALIASES)}); header={list(header)}"
        )

    host_col = _find_col(header, _HOST_ALIASES)
    computer_col = _find_col(header, (_COMPUTER_COL,))
    if host_col is None and computer_col is None:
        raise ValueError(
            f"{path}: sheet {actual_sheet!r} header has neither 'Hostname' nor 'Computer' "
            f"column; header={list(header)}"
        )

    mapping: dict[int, list[str]] = {}
    for row in rows[_HEADER_ROW_INDEX + 1:]:
        group = row[group_col]
        if group is None:
            continue
        try:
            group = int(group)
        except (TypeError, ValueError):
            continue

        if host_col is not None:
            raw_host = row[host_col]
        else:
            raw_computer = row[computer_col]
            raw_host = _bare_hostname(str(raw_computer)) if raw_computer else None

        if not raw_host:
            continue
        hostname = str(raw_host).strip()
        if not hostname or hostname not in inventory.hosts:
            continue

        if hostname not in mapping.setdefault(group, []):
            mapping[group].append(hostname)

    return {g: tuple(hosts) for g, hosts in mapping.items()}


def hosts_for_groups(groups: tuple[int, ...], wave_mapping: dict[int, tuple[str, ...]]) -> tuple[str, ...]:
    hosts: list[str] = []
    for group in groups:
        if group not in wave_mapping:
            raise KeyError(
                f"group {group} has no hosts in the wave mapping — "
                f"check that those hosts appear in the Excel host sheet and in hosts.yaml"
            )
        for host in wave_mapping[group]:
            if host not in hosts:
                hosts.append(host)
    return tuple(hosts)


# ---------------------------------------------------------------------------
# Kept for reference / manual override: YAML-based loading is no longer used
# by the run command.
# ---------------------------------------------------------------------------

def generate_wave_mapping_template(team_steps: list[TeamStep]) -> str:
    groups: set[int] = set()
    for step in team_steps:
        groups.update(step.groups)
    lines = ["# Autogenerated — not needed when running via the Excel host sheet.", "groups:"]
    for group in sorted(groups):
        lines.append(f"  {group}: []")
    return "\n".join(lines) + "\n"
